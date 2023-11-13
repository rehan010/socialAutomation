
from app.models import Wilayas, WilayaPopulation
import logging

from fuzzywuzzy import fuzz
from unidecode import unidecode
import openpyxl
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Process an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the .xlsx file to process')

    def handle(self, *args, **options):
        file_path = options['file_path']

        # Attempt to open the .xlsx file
        try:
            workbook = openpyxl.load_workbook(file_path)
            # Access the sheets and process the data
            sheets = workbook.sheetnames
            for sheet_name in sheets:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Assuming 'Wilaya' is in the first column, 'Population Résidente 2022' in the second column, and so on

                        normalized = round((100 * row[1] / 45500000), 2)
                        wilaya_name = row[0]
                        wilaya_name = unidecode(wilaya_name)
                        matching_wilaya = Wilayas.objects.all()
                        similarity_threshold = 85
                        matching_records = None

                        for wilaya_record in matching_wilaya:
                            wilaya_record_name_normalized = unidecode(wilaya_record.name)
                            similarity_score = fuzz.ratio(wilaya_name, wilaya_record_name_normalized)

                            if ("Algiers" in wilaya_record_name_normalized and "Alger" in wilaya_name) or (
                                    "M'Sila" in wilaya_record_name_normalized and "M'sila" in wilaya_name):
                                similarity_score = 200

                            if similarity_score >= similarity_threshold:
                                matching_records = wilaya_record
                                break

                        logging.log(msg=matching_records, level=1)


                        wilaya = matching_records


                        wilaya_population, created = WilayaPopulation.objects.get_or_create(
                            wilaya=wilaya,
                            defaults={
                                'count': round(row[1], 2),
                                'normalized': normalized,
                                'norm_coeff': round((normalized*10), 2)
                            }
                        )

                        if not created:  # Update values if the record already exists
                            wilaya_population.count = row[1]
                            wilaya_population.normalized = normalized
                            wilaya_population.norm_coeff = normalized*10
                            wilaya_population.save()
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(f"Error processing row: {e}"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error processing file: {e}"))

